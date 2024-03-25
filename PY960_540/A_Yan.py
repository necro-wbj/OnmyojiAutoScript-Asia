#TODO get pic and CUT TO 700X200

YYS_state = 3

DEBUG = 1

###
import pyautogui
import pytesseract
from openpyxl import load_workbook
import jieba
import difflib
import time
import numpy
from PIL import ImageGrab
from PIL import Image
import cv2
import re
import adbutils
import numpy as np
import os

import logging

# 設定logging to utf8 not big5
logging.basicConfig(filename='./pic/Yan/save_pic/ans_pic.txt', level=logging.INFO, filemode='w', encoding='utf-8')
YYS_state = 3
PIC_GET_FLAG = 0 # 0: get pic from screen, 1: get pic from file
MOVE_FLAG = 1 # 0: no move, 1: move
go_firest_move_flag = 1
devices = adbutils.adb.device_list()
TT = devices[1]
YY = devices[0]
def movebystate(go_state):
    global TT
    global YYS_state 
    global move_x
    global move_y
    global YYS_Do_move
    global MOVE_FLAG

    move_x = 147
    move_y = 425
    if MOVE_FLAG:
        if go_state>4:
            print("go_state too big do nothing", go_state)
        elif go_state==0:
            print("go_state too low do nothing", go_state)
        elif (MOVE_FLAG>0) and (move_x>0):
            if YYS_state == 1 or YYS_state == 3:
                if go_state == YYS_state +1:
                    #move right
                    TT.swipe(move_x,move_y,move_x+100,move_y+0,0.7)#move right
                    print("move right")
                    YYS_state = YYS_state +1
            elif YYS_state == 2 or YYS_state == 4:
                if go_state == YYS_state -1:
                    #move right
                    TT.swipe(move_x,move_y,move_x-100,move_y+0,0.7)#move left
                    YYS_state = YYS_state -1
            if YYS_state == 1 or YYS_state == 2:
                if go_state == YYS_state +2:
                    #move right
                    TT.swipe(move_x,move_y,move_x+0,move_y+100,0.45)#move down
                    YYS_state = YYS_state +2
            elif YYS_state == 3 or YYS_state == 4:
                if go_state == YYS_state -2:
                    #move right
                    TT.swipe(move_x,move_y,move_x+0,move_y-100,0.45)#move up
                    YYS_state = YYS_state -2
            if YYS_state == 1 and go_state == 4:
                TT.swipe(move_x,move_y,move_x+100,move_y+0,0.7)#move right
                TT.swipe(move_x,move_y,move_x+0,move_y+100,0.45)#move down
                YYS_state = go_state
            elif YYS_state == 4 and go_state == 1:
                TT.swipe(move_x,move_y,move_x-100,move_y+0,0.7)#move left
                TT.swipe(move_x,move_y,move_x+0,move_y-100,0.45)#move up
                YYS_state = go_state
            elif YYS_state == 2 and go_state == 3:
                TT.swipe(move_x,move_y,move_x+0,move_y+100,0.45)#move down
                TT.swipe(move_x,move_y,move_x-100,move_y+0,0.7)#move left
                YYS_state = go_state
            elif YYS_state == 3 and go_state == 2:
                TT.swipe(move_x,move_y,move_x+100,move_y+0,0.7)#move right
                TT.swipe(move_x,move_y,move_x+0,move_y-100,0.45)#move up
                YYS_state = go_state
            print("now YYS state ",YYS_state)
        print("TODO")
if 1:
    pic_list = ["./pic/Yan/sam_pic/1_1T.png"]
    print(pic_list)
    # get all pic from ./pic/Yan/sam_pic/
else :
    pic_list = []
    dir_list = os.listdir("./pic/Yan/sam/")
    for dir_item in dir_list:
        pic_list.append("./pic/Yan/sam/" + dir_item)
    print(pic_list)

def text_error_handle(text):
    text = text.replace("御魂 狠 ", "禦魂 猙 ")
    text = text.replace("蕙", "薰")
    text = text.replace("蔥", "薰")
    text = text.replace("自", "白")
    text = text.replace("\n","")
    text = text.replace("|","")
    text = text.replace(" ","")
    text = text.replace("「", " ")
    text = text.replace("」", " ")
    text = text.replace("'", " ")
    text = text.replace("《", " ")
    text = text.replace("》", " ")
    text = text.replace("】", "]")
    return text

def cut_pic(pic):
    X_MAX = [164,700]
    Y_MAX = [40,170]
    for y in range(0, pic.shape[0]):
        for x in range(0, pic.shape[1]):
            if y < Y_MAX[0] or y > Y_MAX[1] or x < X_MAX[0] or x > X_MAX[1]:
                pic[y ,x] = [0,0,0]
    return pic

def get_pic(num):
    if PIC_GET_FLAG: # get pic from file
        print("get pic from file")
        # get pic from ./pic/Yan/sam_pic/X_X.png
        pic = cv2.imread(pic_list[num])
        
    else:            # get pic from screen
        print("get pic from screen")
        # get device list from adb
        devices = adbutils.adb.device_list()
        # get device model==SM_S901N
        # for device in devices:
        #     if device.serial == "emulator-5556":
        #         TT = device
        TT = devices[1]
        pic = TT.screenshot()
        pic = cv2.cvtColor(np.array(pic), cv2.COLOR_RGB2BGR)
        #show pic
        # cv2.imshow(f"PIC", pic)
        # cv2.waitKey()
    # only use X_MAX and Y_MAX other place draw black
    pic = cut_pic(pic)
    #save pic to ./pic/Yan/save_pic/X_X.png
    cv2.imwrite("./pic/Yan/save_pic/X_X.png", pic)
    return pic
    

def get_pic_NO_CUT(num):
    if PIC_GET_FLAG: # get pic from file
        print("get pic from file")
        # get pic from ./pic/Yan/sam_pic/X_X.png
        pic = cv2.imread(pic_list[num])
        
    else:            # get pic from screen
        print("get pic from screen")
        # get device list from adb
        devices = adbutils.adb.device_list()
        # get device model==SM_S901N
        # for device in devices:
        #     if device.serial == "emulator-5556":
        #         TT = device
        TT = devices[1]
        pic = TT.screenshot()
        pic = cv2.cvtColor(np.array(pic), cv2.COLOR_RGB2BGR)
        #show pic
        # cv2.imshow(f"PIC", pic)
        # cv2.waitKey()
    #save pic to ./pic/Yan/save_pic/X_X.png
    cv2.imwrite("./pic/Yan/save_pic/X_X.png", pic)
    return pic

Alist=[
"./pic/Yan/A4.png",             #O
"./pic/Yan/A3.png",             #O
"./pic/Yan/A2.png",             #O
"./pic/Yan/A1.png",             #X
"./pic/Yan/A1A.png",            #?
"./pic/Yan/A2A.png",            #X
"./pic/Yan/A3A.png",            #O
"./pic/Yan/A4A.png"             #O
]
color_list = [
    (0, 0, 100),
    (0, 0, 255),
    (0, 255, 255),
    (255, 0, 255),
    (255, 255, 0),
    (100, 0, 100),
    (0, 111, 111),
    (111, 0, 111),
]

template_list = [] 
for Aitem in Alist:
    template = cv2.imread(Aitem)
    template_list.append(template)

#setup excel
wb = load_workbook("./pic/FM/yys.xlsx")
sheet = wb['Sheet1']
my_list = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row+1)]
EXCEL_ANS_list = [sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row+1)]


# here check pic can match A4-A2 and max_val_sum > 2.3
max_val_sum = 0 
max_loclist = []
# TT=YY
# check MOVE_FLAG and time to move
if 1:
    pic = get_pic_NO_CUT(0)
    screenshot = pic
    YanHui_XYLL_N = cv2.imread("./pic/Yan/new_YYL_UI_LiaoXinXi.png")
    image_to_find = YanHui_XYLL_N
    result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
    min_val, max_val_N, min_loc, max_loc = cv2.minMaxLoc(result)
    if max_val_N > 0.3:
        move_x = max_loc[0]-752 +27
        move_y = max_loc[1]-92 +11
    else:
        move_x = 0
        move_y = 0
    move_x = 147
    move_y = 425
    print("new_YYL_UI_LiaoXinXi val",max_val_N)
    YYL_MOVE = cv2.imread("./pic/Yan/YYL_MOVE.png")
    image_to_find = YYL_MOVE
    result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    print("YYL_MOVE val",max_val)
    if max_val > 0.8 and go_firest_move_flag:
        # TT=YY
        TT.swipe(move_x,move_y,move_x+100,move_y+0,0.2)#move right
        TT.swipe(move_x,move_y,move_x-100,move_y+0,0.4)#move left
        TT.swipe(move_x,move_y,move_x+0,move_y-100,0.8)#move up
        TT.swipe(move_x,move_y,move_x-100,move_y+0,2)#move left
        TT.swipe(move_x,move_y,move_x+100,move_y+0,0.8)#move right


    timenow = time.strftime('%m%d_%H%M%S',time.localtime(time.time()))
    cv2.imwrite("./pic/Yan/save_pic/yan_start_"+str(timenow)+".png", screenshot)


# movebystate(4)
# movebystate(2)
# movebystate(1)
# movebystate(3)
all_while_count = 100000
# if 1:
while all_while_count:
    #print all_while_count
    all_while_count = all_while_count -1
    if all_while_count == 0:
        break
    print("all_while_count:",all_while_count)
    # for pic_item in range(0,1):
    for pic_item in pic_list:
        # if pic_item == pic_list[1]:
        #     break
        if DEBUG:
            DEBUG = DEBUG
            # print(pic_item)
        #init pic loc list
        max_loclist = []
        # pic = cv2.imread(pic_item)
        # pic = cut_pic(pic)
        pic = get_pic(pic_list.index(pic_item))
        # pic = get_pic(0)
        screenshot = pic
        cv2draw = screenshot
        max_val_sum = 0
        for count in range(0,3):
            template = template_list[count]
            # match picture in screen
            res = cv2.matchTemplate(screenshot, template, cv2.TM_CCOEFF_NORMED)
            min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
            max_val_sum = max_val_sum + max_val
            max_loclist.append(max_loc)
            if DEBUG:
                # print(f"[{count}]:{pic_item},{max_val},{max_loc}")
                # draw rectangle in cv2 window
                cv2draw = cv2.rectangle(cv2draw, max_loc, (max_loc[0] + 100, max_loc[1] + 20), color_list[count], 2)
        print(max_val_sum)
        if max_val_sum > 2.2:
            print("max_val_sum > 2.2")
            all_while_count = 100
        else:
            print("max_val_sum < 2.2")
        if DEBUG:
            DEBUG = DEBUG
            # cv2.imshow(f"PIC", cv2draw)
            # cv2.waitKey()
        GO_NEXT_FLAG = 0
        #get ans pic of A4-A2 and save it in one pic named ./pic/Yan/save_pic/ans_pic.png
        if max_val_sum > 2.2:
            # cv2.imshow(f"PIC", pic)
            # cv2.waitKey()
            # cut pic from max_loclist[0]
            pic0 = pic[max_loclist[0][1]:max_loclist[0][1]+20,max_loclist[0][0]:max_loclist[0][0]+100]
            GO_NEXT_FLAG = 1
            # cv2.imshow(f"pic0", pic0)
            # cv2.waitKey()
        else:
            
            # with open("./pic/Yan/save_pic/ans_pic.txt", "w") as f:
            logging.info(pic_item)
            logging.info("max_val_sum < 2.3")
        # find ./pic/Yan/A4.png and black its right side
        if GO_NEXT_FLAG:
            GO_NEXT_FLAG = 0
            template = template_list[0] # 0 is A4
            # pic = cv2.imread(pic_item)
            pic = get_pic(pic_list.index(pic_item))
            # pic = get_pic(0)

            screenshot = pic
            # res = cv2.matchTemplate(screenshot, template, cv2.TM_CCOEFF_NORMED)
            # min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
            A4max_loc = max_loclist[0] #A4
            A4_X = A4max_loc[0]
            A4_Y = A4max_loc[1]
            A3max_loc = max_loclist[1] #A3
            A3_X = A3max_loc[0]
            A3_Y = A3max_loc[1]
            A2max_loc = max_loclist[2] #A3
            A2_X = A2max_loc[0]
            A2_Y = A2max_loc[1]
            # def done AX_X_Y
            GO_NEXT_FLAG = 1
            # max_loc[0] + 100 is right side of A4 set it to black
            # max_loclist.append(max_loc)
            check_A_Q = 1
            if 0:
                pic = get_pic(pic_list.index(pic_item))
                # pic = get_pic(0)
                AllQandA = pic
                # pic = cv2.imread(pic_item)
                # AllQandA = cut_pic(pic)
                for y in range(0, pic.shape[0]):
                    for x in range(0, pic.shape[1]):
                        if  x > (A4max_loc[0] + 100) or x < (A3max_loc[0] - 10):
                            AllQandA[y ,x] = [0,0,0]
                cv2.imshow(f"pic0", AllQandA)
                cv2.waitKey()
        if GO_NEXT_FLAG:
            GO_NEXT_FLAG = 0 #check_A = 1
            pic = get_pic(pic_list.index(pic_item))
            # cv2.imshow(f"pic0", pic)
            # cv2.waitKey()
            # pic = get_pic(0)

            #handle A1
            # creat AllA as pic copy
            AllA = pic
            # pic = cv2.imread(pic_item)
            # AllA = cut_pic(pic)
            #cut A1 x use A3_X+10:A4_X,y use A2_Y:A2_Y+20
            #print A2max_loc,A3max_loc,A4max_loc for debug
            print(A2max_loc,A3max_loc,A4max_loc)
            # A1_img = AllA
            A1_img = AllA[A2max_loc[1]:A2max_loc[1]+20,A3max_loc[0]+10:A4max_loc[0]]
            #handle A2
            A2_img = AllA[A2max_loc[1]:A2max_loc[1]+20,A2max_loc[0]+15:A2max_loc[0]+100]
            #handle A3
            A3_img = AllA[A3max_loc[1]:A3max_loc[1]+20,A3max_loc[0]+17:A4max_loc[0]-5]
            #handle A4
            A4_img = AllA[A4max_loc[1]:A4max_loc[1]+20,A4max_loc[0]+17:A4max_loc[0]+100]

            A_list = [A1_img,A2_img,A3_img,A4_img]
            #check A*_img is OK
            flag_A_PIC_FAIL = 0
            for A_item in A_list:
                if A_item.size == 0:
                    logging.info(pic_item)
                    logging.info("A_item.size == 0")
                    flag_A_PIC_FAIL = flag_A_PIC_FAIL +1
                    # with open("./pic/Yan/save_pic/ans_pic.txt", "w") as f:
                    #     f.write(pic_item)
                    #     f.write("A_item.size == 0")
                    print("A_item.size == 0")
                    continue
            if flag_A_PIC_FAIL:
                continue
            #combine A1-A4 to one pic
            A_img = np.hstack((A1_img,A2_img,A3_img,A4_img))
            #save pic to ./pic/Yan/save_pic/A_pic.png
            cv2.imwrite("./pic/Yan/save_pic/A_pic.png", A_img)
            GO_NEXT_FLAG = 1



                # cv2.imshow(f"pic0", A4_img)
                # cv2.waitKey()
        if GO_NEXT_FLAG:#check_Q = 1
            GO_NEXT_FLAG = 0
            pic = get_pic(pic_list.index(pic_item))
            # pic = get_pic(0)
            AllQ = pic
            # pic = cv2.imread(pic_item)
            # AllQ = cut_pic(pic)
            for y in range(0, pic.shape[0]):
                for x in range(0, pic.shape[1]):
                    if  x > (A4max_loc[0] + 100) or x < (A3max_loc[0] - 10) or (y > (A2max_loc[1] -10 )) or y < 40:
                        AllQ[y ,x] = [0,0,0]
            #cut AllQ from((A3max_loc[0] - 10),40) to ((A4max_loc[0] + 100),(A2max_loc[1] -10 ))
            AllQ=AllQ[40:A2max_loc[1] -10,(A3max_loc[0] - 10):A4max_loc[0] + 100]
            # check AllQ pic is OK no show below error
            # cv2.error: OpenCV(4.7.0) D:\a\opencv-python\opencv-python\opencv\modules\imgcodecs\src\loadsave.cpp:783: error: (-215:Assertion failed) !_img.empty() in function 'cv::imwrite'
            if AllQ.size == 0:
                logging.info(pic_item)
                logging.info("AllQ.size == 0")
                # with open("./pic/Yan/save_pic/ans_pic.txt", "w") as f:

                #     f.write(pic_item)
                #     f.write("AllQ.size == 0")
                #     print("AllQ.size == 0")
                continue
            #save pic to ./pic/Yan/save_pic/ans_pic.png

            cv2.imwrite("./pic/Yan/save_pic/ans_pic.png", AllQ)
            #save piv with time stamp
            cv2.imwrite("./pic/Yan/save_pic/Q_pic" + str(time.time()) + ".png", AllQ)
            GO_NEXT_FLAG = 1
        if GO_NEXT_FLAG:
            GO_NEXT_FLAG = 0#            ORC_FLAG = 1

            # AllQ ORC
            #in here psm 6>1> other
            confignow = ("-l chi_tra --oem 3 --psm 6")
            print("psm 6")
            logging.info("psm 6 Q:")
            text = pytesseract.image_to_string(AllQ,config=confignow)
            text = text_error_handle(text)
            print(text)
            #save text to ./pic/Yan/save_pic/ans_pic.txt
            logging.info(text)
            QQ_text = text
            # with open("./pic/Yan/save_pic/ans_pic.txt", "w") as f:
            #     f.write(pic_item)
            #     f.write(text)
            

            GO_NEXT_FLAG = 1
        if GO_NEXT_FLAG:
            GO_NEXT_FLAG = 0 #here compare text with excel
            
            sent2sim= []
            Question_text = QQ_text
            #Question_text 分词
            
            # Question_text_lcut = jieba.lcut(Question_text)
            Question_text_lcut = set(list(str(Question_text)))
            GO_NEXT_FLAG = 1
            if Question_text_lcut:
                GO_NEXT_FLAG = 1
            if GO_NEXT_FLAG:
                for sent2 in my_list:
                    # EXCEL 分词
                    # sent2_cut = jieba.lcut(sent2)
                    set2 = set(list(str(sent2)))

                    # 计算相似程度
                    similarity = len(Question_text_lcut & set2) / len(Question_text_lcut | set2)
                    # similarity = difflib.SequenceMatcher(None, Question_text_lcut, sent2_cut).quick_ratio()
                    #add SIM in to list sent2sim
                    sent2sim.append(similarity)
                # print("####2")
            similarity_max_value = 0
            if GO_NEXT_FLAG:
                FIND_ANS_list =[]
                similarity_max_value = max(sent2sim)
                if similarity_max_value < 0.1:
                    GO_NEXT_FLAG=0
            if GO_NEXT_FLAG:
                # print("similarity_max_value",similarity_max_value)
                # # this is find index of MAX similarity sent2sim.index(similarity_max_value)
                similarity_max_index = sent2sim.index(similarity_max_value)

                for simii in range(len(sent2sim)):
                    if similarity_max_value == sent2sim[simii]:
                        # sent2_cut = jieba.lcut(my_list[simii])
                        # print("index:",simii)
                        # print("my_list:",my_list[simii])
                        # print("sent2_cut:",sent2_cut)
                        # print("EXCEL_A:",EXCEL_ANS_list[simii])
                        print("EXCEL line[",simii,"],Q:[",my_list[simii],"],A:[",EXCEL_ANS_list[simii],"]")

                        FIND_ANS_list.append(EXCEL_ANS_list[simii])
            if my_list[0] is None:
                GO_NEXT_FLAG=0
        if GO_NEXT_FLAG: #TODO feature auto update ANS w as excel ans
            FLAG_ANS_ONLY_NUM = 0
            if len(FIND_ANS_list) == 1:
                if isinstance(FIND_ANS_list[0],int):
                    print("ans only number")
                    FLAG_ANS_ONLY_NUM = 1
                    
                    if FIND_ANS_list[0] <10 :
                        print("ans number <10 set w is small")
                        ANS_match_loc_w = 10
                else:
                    ANS_match_loc_w = len(FIND_ANS_list[0]) *13
            else:
                #find max len in FIND_ANS_list item
                max_ANS_match_loc_w = 0
                for FIND_ANS_list_item in FIND_ANS_list:
                    if isinstance(FIND_ANS_list_item,int):
                        if FIND_ANS_list[0] <10 :
                            print("ans number <10 set w is small")
                            if max_ANS_match_loc_w < 10:
                                max_ANS_match_loc_w = 10
                        else:
                            
                            if max_ANS_match_loc_w < 20:
                                max_ANS_match_loc_w = 10
                    else:
                        if max_ANS_match_loc_w < len(FIND_ANS_list_item) *13:
                            max_ANS_match_loc_w = len(FIND_ANS_list_item) *13
                ANS_match_loc_w = max_ANS_match_loc_w
            # cut A1_img W if > ANS_match_loc_w
            if A1_img.shape[1] > ANS_match_loc_w:
                A1_img = A1_img[0:A1_img.shape[0],0:ANS_match_loc_w]
            # cut A2_img W if > ANS_match_loc_w
            if A2_img.shape[1] > ANS_match_loc_w:
                A2_img = A2_img[0:A2_img.shape[0],0:ANS_match_loc_w]
            # cut A3_img W if > ANS_match_loc_w
            if A3_img.shape[1] > ANS_match_loc_w:
                A3_img = A3_img[0:A3_img.shape[0],0:ANS_match_loc_w]
            # cut A4_img W if > ANS_match_loc_w
            if A4_img.shape[1] > ANS_match_loc_w:
                A4_img = A4_img[0:A4_img.shape[0],0:ANS_match_loc_w]
        if GO_NEXT_FLAG:
            
            #trans A1_img to text
            #about psm 1.6.11 is good for A
            confignow = ("-l chi_tra --oem 3 --psm 6")
            if FLAG_ANS_ONLY_NUM:
                confignow = ("-l eng --oem 3 --psm 7")
            text = pytesseract.image_to_string(A1_img,config=confignow)
            text = text_error_handle(text)
            logging.info("A1:")
            logging.info(text)
            print("A1:",text)
            A1_text = text

            #trans A2_img to text
            text = pytesseract.image_to_string(A2_img,config=confignow)
            text = text_error_handle(text)
            logging.info("A2:")
            logging.info(text)
            print("A2:",text)
            A2_text = text

            #trans A3_img to text
            text = pytesseract.image_to_string(A3_img,config=confignow)
            text = text_error_handle(text)
            logging.info("A3:")
            logging.info(text)
            print("A3:",text)
            A3_text = text

            #trans A4_img to text
            text = pytesseract.image_to_string(A4_img,config=confignow)
            text = text_error_handle(text)
            logging.info("A4:")
            logging.info(text)
            print("A4:",text)
            A4_text = text
        if GO_NEXT_FLAG:
            YYS_ANS_list = [A1_text,A2_text,A3_text,A4_text]
            for FIND_ANS_item in FIND_ANS_list:
                FIND_ANS_item_cut = set(list(str(FIND_ANS_item)))
                if isinstance(FIND_ANS_list[0],int):
                    print("ans only number")
            similarity_max = 0
            th_hold = 0.3
            YYS_possible_ANS_list = []
            if 0:#isinstance(FIND_ANS_list[0], (int,float)):
                # print("ANS NO CHINESE")
                for FIND_ANS_item in FIND_ANS_list:
                    for YYS_ANS_item in YYS_ANS_list:
                        if FIND_ANS_item == YYS_ANS_item:
                            print("maybe index [",YYS_ANS_list.index(YYS_ANS_item)+1,"],ans is [",FIND_ANS_item,"]")
            else:
                for FIND_ANS_item in FIND_ANS_list:
                    # FIND_ANS_item_cut = jieba.lcut(FIND_ANS_item)
                    FIND_ANS_item_cut = set(list(str(FIND_ANS_item)))
                    for YYS_ANS_item in YYS_ANS_list:
                        # if len(FIND_ANS_item)> len(YYS_ANS_item):
                        #     # YYS_ANS_list[YYS_ANS_list.index(YYS_ANS_item)] = YYS_ANS_item +"xo"
                        #     # YYS_ANS_item = YYS_ANS_item +"xo"
                        #     YYS_ANS_item = YYS_ANS_item
                        # YYS_ANS_item_cut = jieba.lcut(YYS_ANS_item)
                        YYS_ANS_item_cut = set(list(str(YYS_ANS_item)))
                        similarity = len(FIND_ANS_item_cut & YYS_ANS_item_cut) / len(FIND_ANS_item_cut | YYS_ANS_item_cut)
                        # similarity = difflib.SequenceMatcher(None, FIND_ANS_item_cut, YYS_ANS_item_cut).quick_ratio()
                        # print(FIND_ANS_item_cut,";",YYS_ANS_item_cut,"S:",similarity)
                        if similarity_max < similarity:
                            similarity_max = similarity
                            YYS_possible_ANS_list = []
                        if similarity_max == similarity:
                            YYS_possible_ANS_list.append(YYS_ANS_item)
                print("possible ans ",YYS_possible_ANS_list)
                if len(YYS_possible_ANS_list) == 1:
                    YYS_ANS_item=YYS_possible_ANS_list[0]
                    print("choose only one",YYS_ANS_list.index(YYS_ANS_item)+1)
                    movebystate(YYS_ANS_list.index(YYS_ANS_item)+1)



ADicon = cv2.imread("./pic/Yan/android_icon.png")
image_to_find = cv2.imread("./pic/Yan/yan_question_end.png") #70
image_to_find = cv2.imread("./pic/Yan/yan_question_R_bracket1.png") #40
image_to_find = cv2.imread("./pic/Yan/yan_question_LT.png") #65
A1_img=cv2.imread("./pic/Yan/A1.png")
A2_img=cv2.imread("./pic/Yan/A2.png")
A3_img=cv2.imread("./pic/Yan/A3.png")
A4_img=cv2.imread("./pic/Yan/A4.png")
A1_img=cv2.imread("./pic/Yan/A1A.png")
A2_img=cv2.imread("./pic/Yan/A2A.png")
A3_img=cv2.imread("./pic/Yan/A3A.png")
A4_img=cv2.imread("./pic/Yan/A4A.png")
YanHui_XYLL_N = cv2.imread("./pic/Yan/1666177652310.png")
YanHui_LiaoXinXi = cv2.imread("./pic/Yan/new_YYL_UI_LiaoXinXi.png")
YanHui_XYLL_N = cv2.imread("./pic/Yan/1673188292135.png")
YanHui_XYLL_N = cv2.imread("./pic/Yan/new_YYL_UI_LiaoXinXi.png")
YanHui_XYLL_N = cv2.imread("./pic/Yan/new_YYL_UI_LiaoXinXi.png")
YanHui_XYLL_Y = cv2.imread("./pic/Yan/1673188475535.png")
image_to_find = cv2.imread("./pic/Yan/yan_question_end.png") #70
image_to_find = cv2.imread("./pic/Yan/yan_question_R_bracket1.png") #40
image_to_find = cv2.imread("./pic/Yan/yan_question_R_bracket1.png") #65
image_to_find = cv2.imread("./pic/Yan/yan_question_LT.png") #65
image_to_find = cv2.imread("./pic/Yan/yan_question_LT.png") #65
A1_img=cv2.imread("./pic/Yan/A1.png")
A2_img=cv2.imread("./pic/Yan/A2.png")
A3_img=cv2.imread("./pic/Yan/A3.png")
A4_img=cv2.imread("./pic/Yan/A4.png")
A1_img=cv2.imread("./pic/Yan/A1A.png")
A2_img=cv2.imread("./pic/Yan/A2A.png")
A3_img=cv2.imread("./pic/Yan/A3A.png")
A4_img=cv2.imread("./pic/Yan/A4A.png")

# "./pic/Yan/A1.png","./pic/Yan/sam_pic/1_1T.png",    0.9110     X
# "./pic/Yan/A1.png","./pic/Yan/sam_pic/1_2Y.png",    0.9319     o
# "./pic/Yan/A1.png","./pic/Yan/sam_pic/3_2T.png",    0.9534     x
# "./pic/Yan/A1.png","./pic/Yan/sam_pic/3_1Y.png",    0.9828     o
# "./pic/Yan/A1.png","./pic/Yan/sam_pic/4_1Y.png",    0.9962     o
# "./pic/Yan/A1.png","./pic/Yan/sam_pic/4_2T.png",    0.9491     o
     
# "./pic/Yan/A2.png","./pic/Yan/sam_pic/1_1T.png",    0.8386     o
# "./pic/Yan/A2.png","./pic/Yan/sam_pic/1_2Y.png",    0.9763     o
# "./pic/Yan/A2.png","./pic/Yan/sam_pic/3_2T.png",    0.9026     
# "./pic/Yan/A2.png","./pic/Yan/sam_pic/3_1Y.png",    0.9955     
# "./pic/Yan/A2.png","./pic/Yan/sam_pic/4_1Y.png",    0.9582     
# "./pic/Yan/A2.png","./pic/Yan/sam_pic/4_2T.png",    0.9203     
     
# "./pic/Yan/A3.png","./pic/Yan/sam_pic/1_1T.png",    0.8082     
# "./pic/Yan/A3.png","./pic/Yan/sam_pic/1_2Y.png",    0.9833     
# "./pic/Yan/A3.png","./pic/Yan/sam_pic/3_2T.png",    0.7675     
# "./pic/Yan/A3.png","./pic/Yan/sam_pic/3_1Y.png",    0.8093     
# "./pic/Yan/A3.png","./pic/Yan/sam_pic/4_1Y.png",    0.9948     
# "./pic/Yan/A3.png","./pic/Yan/sam_pic/4_2T.png",    0.8931     
     
# "./pic/Yan/A4.png","./pic/Yan/sam_pic/1_1T.png",    0.7723     
# "./pic/Yan/A4.png","./pic/Yan/sam_pic/1_2Y.png",    0.9666     
# "./pic/Yan/A4.png","./pic/Yan/sam_pic/3_2T.png",    0.8499     
# "./pic/Yan/A4.png","./pic/Yan/sam_pic/3_1Y.png",    0.9930     
# "./pic/Yan/A4.png","./pic/Yan/sam_pic/4_1Y.png",    0.8448     
# "./pic/Yan/A4.png","./pic/Yan/sam_pic/4_2T.png",    0.7783     
     
# "./pic/Yan/A1A.png","./pic/Yan/sam_pic/1_1T.png",   0.9627     
# "./pic/Yan/A1A.png","./pic/Yan/sam_pic/1_2Y.png",   0.9114     
# "./pic/Yan/A1A.png","./pic/Yan/sam_pic/3_2T.png",   0.9120     
# "./pic/Yan/A1A.png","./pic/Yan/sam_pic/3_1Y.png",   0.8615     
# "./pic/Yan/A1A.png","./pic/Yan/sam_pic/4_1Y.png",   0.8475     
# "./pic/Yan/A1A.png","./pic/Yan/sam_pic/4_2T.png",   0.9276     
     
# "./pic/Yan/A2A.png","./pic/Yan/sam_pic/1_1T.png",   0.8501     
# "./pic/Yan/A2A.png","./pic/Yan/sam_pic/1_2Y.png",   0.7478     
# "./pic/Yan/A2A.png","./pic/Yan/sam_pic/3_2T.png",   0.8241     
# "./pic/Yan/A2A.png","./pic/Yan/sam_pic/3_1Y.png",   0.7443     
# "./pic/Yan/A2A.png","./pic/Yan/sam_pic/4_1Y.png",   0.7446     
# "./pic/Yan/A2A.png","./pic/Yan/sam_pic/4_2T.png",   0.8208     
     
# "./pic/Yan/A3A.png","./pic/Yan/sam_pic/1_1T.png",   0.8287     
# "./pic/Yan/A3A.png","./pic/Yan/sam_pic/1_2Y.png",   0.7647     
# "./pic/Yan/A3A.png","./pic/Yan/sam_pic/3_2T.png",   0.9995     
# "./pic/Yan/A3A.png","./pic/Yan/sam_pic/3_1Y.png",   0.8539     
# "./pic/Yan/A3A.png","./pic/Yan/sam_pic/4_1Y.png",   0.7683     
# "./pic/Yan/A3A.png","./pic/Yan/sam_pic/4_2T.png",   0.8382     
     
# "./pic/Yan/A4A.png","./pic/Yan/sam_pic/1_1T.png",   0.8661     
# "./pic/Yan/A4A.png","./pic/Yan/sam_pic/1_2Y.png",   0.7851     
# "./pic/Yan/A4A.png","./pic/Yan/sam_pic/3_2T.png",   0.8743     
# "./pic/Yan/A4A.png","./pic/Yan/sam_pic/3_1Y.png",   0.8048     
# "./pic/Yan/A4A.png","./pic/Yan/sam_pic/4_1Y.png",   0.8755     
# "./pic/Yan/A4A.png","./pic/Yan/sam_pic/4_2T.png",   0.9733     
     