#TODO get pic and CUT TO 700X200

YYS_state = 3

DEBUG = 0

###
import pyautogui
import pytesseract
from openpyxl import load_workbook
import jieba
import difflib
import time
import numpy
from PIL import ImageGrab
from PIL import Image
import cv2
import re


import adbutils
import numpy as np

flag_FM_ANS = 1
global YYS_Do_move
YYS_Do_move = 0
if DEBUG == 1:
    YYS_Do_move = 0
movefromtingyuan = 1
global move_x
global move_y
YYS_Q_Num = 0
OLDQuestion_numbers = 0
BaseOnStateMoveAns = 1
test = 0
if test:
    YYS_Do_move = 0
    YYS_state = 3
# jieba.load_userdict("userdict.txt")

BaseOnStateMoveMatrix = [[-1,0,-1,0],[0,0,0,0],[-1,0,0,0],[0,0,0,0]]
def get_screenshot():
    if DEBUG == 1:
        screen = cv2.imread("./SC2.png")
        # screen = cv2.cvtColor(numpy.array(screen), cv2.COLOR_RGB2BGR)
        return screen
    else: 
        devices = adbutils.adb.device_list()
        device = devices[0]
        screen = device.screenshot()

        screen = cv2.cvtColor(np.array(screen), cv2.COLOR_RGB2BGR)
        return screen

def text_error_handle(text):
    text = text.replace("御魂 狠 ", "禦魂 猙 ")
    text = text.replace("蕙", "薰")
    text = text.replace("蔥", "薰")
    text = text.replace("自", "白")
    text = text.replace("\n","")
    text = text.replace("|","")
    text = text.replace(" ","")
    text = text.replace("「", " ")
    text = text.replace("」", " ")
    text = text.replace("'", " ")
    text = text.replace("《", " ")
    text = text.replace("》", " ")
    return text

devices = adbutils.adb.device_list()
def movewhere(X,Y,XN,YN,T):
    global YYS_Do_move
    move_x = 147
    move_y = 425
    if YYS_Do_move:
        for device in devices:
            device.swipe(move_x,move_y,move_x+XN,move_y+YN,0.7)
        # pyautogui.mouseDown(X, Y , button='left')
        # time.sleep(0.1)
        # pyautogui.moveTo(X+XN, Y-YN)
        # time.sleep(T)
        # pyautogui.mouseUp(X+XN, Y-YN , button='left')
        # time.sleep(0.5)


def movebystate(go_state):
    global YYS_state 
    global move_x
    global move_y
    global YYS_Do_move
    if go_state>4:
        print("go_state too big do nothing", go_state)
    elif go_state==0:
        print("go_state too low do nothing", go_state)
    elif (YYS_Do_move>0) and (move_x>0):
        if YYS_state == 1 or YYS_state == 3:
            if go_state == YYS_state +1:
                #move right
                movewhere(move_x,move_y,100,0,0.4)#move right
                YYS_state = YYS_state +1
        elif YYS_state == 2 or YYS_state == 4:
            if go_state == YYS_state -1:
                #move right
                movewhere(move_x,move_y,-100,0,0.4)#move left
                YYS_state = YYS_state -1
        if YYS_state == 1 or YYS_state == 2:
            if go_state == YYS_state +2:
                #move right
                movewhere(move_x,move_y,0,100,0.4)#move down
                YYS_state = YYS_state +2
        elif YYS_state == 3 or YYS_state == 4:
            if go_state == YYS_state -2:
                #move right
                movewhere(move_x,move_y,0,-100,0.4)#move up
                
                movewhere(move_x,move_y,0,-100,0.6)#move up
                YYS_state = YYS_state -2
        if YYS_state == 1 and go_state == 4:
            movewhere(move_x,move_y,100,0,0.4)#move right
            movewhere(move_x,move_y,0,100,0.4)#move down
            YYS_state = go_state
        elif YYS_state == 4 and go_state == 1:
            movewhere(move_x,move_y,-100,0,0.4)#move left
            movewhere(move_x,move_y,0,-100,0.4)#move up
            YYS_state = go_state
        elif YYS_state == 2 and go_state == 3:
            movewhere(move_x,move_y,0,100,0.4)#move down
            movewhere(move_x,move_y,-100,0,0.4)#move left
            YYS_state = go_state
        elif YYS_state == 3 and go_state == 2:
            movewhere(move_x,move_y,100,0,0.4)#move right
            movewhere(move_x,move_y,0,-100,0.4)#move up
            YYS_state = go_state
        print("now YYS state ",YYS_state)
#### find YYS Ting
# screenshot = get_screenshot()
# SS_android = screenshot.crop((0, 0, (screenshot.width), 500))
# SS_android = cv2.cvtColor(numpy.array(SS_android), cv2.COLOR_RGB2BGR)
# ADicon = cv2.imread("./pic/Yan/android_icon.png")
# image_to_find = ADicon
# result = cv2.matchTemplate(SS_android, image_to_find, cv2.TM_CCOEFF_NORMED)
# min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)

# YYS_TT_loc = [max_loc[0],max_loc[1],max_loc[0]+960+30,max_loc[1]+540+30]
YYS_TT_loc=0

def TT_LOC_XP(xpercent):
    XP = YYS_TT_loc[2] - YYS_TT_loc[0]
    XP = int(XP * xpercent /100)
    return XP #+YYS_TT_loc[0]
def TT_LOC_YP(ypercent):
    YP = YYS_TT_loc[3] - YYS_TT_loc[1]
    YP = int(YP * ypercent /100)
    return YP #+YYS_TT_loc[1]


if 0: #TODO:change to adb not pyautogui 
    for count in range(3):
        screenshot = get_screenshot()

        YanHui_XYLL_N = cv2.imread("./pic/Yan/1666177652310.png")
        image_to_find = YanHui_XYLL_N
        result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        x = max_loc[0]+23
        y = max_loc[1]+25
        print("ting yuan go YinYangLiao ",max_val)
        if max_val>0.9:
            pyautogui.click(x,y,button='left')
            time.sleep(5)
        YanHui_LiaoXinXi = cv2.imread("./pic/Yan/new_YYL_UI_LiaoXinXi.png")
        image_to_find = YanHui_LiaoXinXi
        result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
        min_val, max_val_N, min_loc, max_loc = cv2.minMaxLoc(result)
        if max_val_N > 0.5:
            break

if test == 0:
    for countmove in range(1):
        screenshot = get_screenshot()

        YanHui_XYLL_N = cv2.imread("./pic/Yan/1673188292135.png")
        image_to_find = YanHui_XYLL_N
        result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        x = max_loc[0]+23
        y = max_loc[1]+25
        print("yan_move controller",max_val)
        if max_val<0.85:
            print("yan_move NO ",max_val)
            movefromtingyuan = 0
        else:
            print("yan_move enough",max_val)
            movefromtingyuan = 1
            time.sleep(5)
            break
        

if 1:
    screenshot = get_screenshot()
    # screenshot = screenshot.crop((0, 0, (screenshot.width), 700))
    # screenshot = cv2.cvtColor(numpy.array(screenshot), cv2.COLOR_RGB2BGR)

    YanHui_XYLL_N = cv2.imread("./pic/Yan/new_YYL_UI_LiaoXinXi.png")
    image_to_find = YanHui_XYLL_N
    result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
    min_val, max_val_N, min_loc, max_loc = cv2.minMaxLoc(result)
    if max_val_N > 0.5:
        move_x = max_loc[0]-752 +27
        move_y = max_loc[1]-92 +11
    else:
        move_x = 0
        move_y = 0
    print("new_YYL_UI_LiaoXinXi val",max_val)
    if movefromtingyuan and YYS_Do_move:
        for count in range(1):
            movewhere(move_x,move_y,100,0,0.2)#move right
        for count in range(1):#3
            movewhere(move_x,move_y,-120,0,0.5)#move left
        for count in range(1):
            movewhere(move_x,move_y,0,-100,0.6)#move up
        for count in range(3):#3
            movewhere(move_x,move_y,-120,0,2)#move left
        #move to YYL Left
        for count in range(1):
            movewhere(move_x,move_y,100,0,0.6)#move right
        #now at 3
        YYS_state = 3
    

    timenow = time.strftime('%m%d_%H%M%S',time.localtime(time.time()))
    cv2.imwrite("yan_start_"+str(timenow)+".png", screenshot)

go_next_flag = 1

YanHui_error_count=0 
Question_numbersOLD = 0
Question_retry_count = 0
while 1:
    # break
    

    screenshot = get_screenshot()
    
    # cv2.imshow("screenshot",screenshot)
    # cv2.waitKey(0)
    # screenshot = screenshot.crop((0, 0, (screenshot.width), 700))
    # screenshot = cv2.cvtColor(numpy.array(screenshot), cv2.COLOR_RGB2BGR)

    YanHui_XYLL_N = cv2.imread("./pic/Yan/new_YYL_UI_LiaoXinXi.png")
    image_to_find = YanHui_XYLL_N
    result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
    min_val, max_val_N, min_loc, max_loc = cv2.minMaxLoc(result)
    # print("max_val_N",max_val_N)

    # if max_val_N>0.7:
    #     pyautogui.moveTo(max_loc[0], max_loc[1])

    YanHui_XYLL_Y = cv2.imread("./pic/Yan/1673188475535.png")
    image_to_find = YanHui_XYLL_Y
    result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
    min_val, max_val_Y, min_loc, max_loc = cv2.minMaxLoc(result)
    # print("max_val_Y",max_val_Y)

    if (max_val_Y>0.7) or (max_val_N>0.7):
        # print("find YanHui_XYLL")
        YanHui_error_count = 0
        go_next_flag = 1
    else:
        YanHui_error_count = YanHui_error_count + 1
        print("NO YanHui count=",YanHui_error_count)
    if YanHui_error_count>300:
        print("NO YanHui ERROR STOP this")
        timenow = time.strftime('%m%d_%H%M%S',time.localtime(time.time()))
        cv2.imwrite("yan_no_"+str(timenow)+".png", screenshot)
        break
    RT_loc = 0
    Question_text = ""
    Question_numbers = 0
    if go_next_flag:
        #get Q
        go_next_flag = 0

        # screenshot = screenshotTT()
        screenshot = get_screenshot()
        screenshotPIL = screenshot
        # screenshot = cv2.cvtColor(numpy.array(screenshot), cv2.COLOR_RGB2BGR)

        
        ##find QQQ
        yan_question_Q = screenshotPIL
        # yan_question_Q = yan_question_Q.crop((yan_question_LT_x1, yan_question_LT_y1, yan_question_LT_x2, yan_question_LT_y3))
        image_to_find = cv2.imread("./pic/Yan/yan_question_end1.png") #70
        
        # yan_question_Q = cv2.cvtColor(numpy.array(yan_question_Q), cv2.COLOR_RGB2BGR)
        result = cv2.matchTemplate(yan_question_Q, image_to_find, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        x1 = 0+max_loc[0]
        x2 = 0+max_loc[0]+9
        y1 = 0+max_loc[1]
        y2 = 0+max_loc[1]+15
        screenshot = cv2.rectangle(screenshot, (x1,y1), (x2,y2), (120,0,120), 1)

        x1 = -230+max_loc[0]
        x2 = +230+max_loc[0]+9
        y1 = -40+max_loc[1]
        y2 = 0+max_loc[1]+15
        screenshot = cv2.rectangle(screenshot, (x1,y1), (x2,y2), (120,0,120), 1)
        
        yan_question_QE = screenshotPIL
        # yan_question_QE = yan_question_QE.crop((x1, y1, x2, y2))
        # yan_question_QE = cv2.cvtColor(numpy.array(yan_question_QE), cv2.COLOR_RGB2BGR)
        yan_question_QE = yan_question_QE[y1:y2,x1:x2]
        yan_question_QE_x1 = x1
        yan_question_QE_y1 = y1
        # cv2.imwrite("yan_question_QE.png", yan_question_QE)
        print("yan_question_QE",max_val)
        if max_val>0.35:
            go_next_flag = 1
            print("yan_question_end HIGH",max_val)
        else: 
            go_next_flag = 0
            print("yan_question_end too low",max_val)
    if go_next_flag:
        image_to_find = cv2.imread("./pic/Yan/yan_question_R_bracket1.png") #40
        
        image_to_find = image_to_find

    if go_next_flag:
        image_to_find = cv2.imread("./pic/Yan/yan_question_R_bracket1.png") #65
        # error: (-215:Assertion failed) _img.size().height <= _templ.size().height && _img.size().width <= _templ.size().width in function 'cv::matchTemplate'
        # if yan_question_QE image size <= image_to_find image size go_next_flag=0
        if yan_question_QE.shape[:2][0] <= image_to_find.shape[:2][0] or yan_question_QE.shape[:2][1] <= image_to_find.shape[:2][1]:
            go_next_flag = 0
            print("ERR 617",max_val)
        
    if go_next_flag:
        result = cv2.matchTemplate(yan_question_QE, image_to_find, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        # print("]",max_val)
        
        x1 = max_loc[0]+yan_question_QE_x1 + 0
        x2 = max_loc[0]+yan_question_QE_x1 + 23
        y1 = max_loc[1]+yan_question_QE_y1 + 0
        y2 = max_loc[1]+yan_question_QE_y1 + 15
        screenshot = cv2.rectangle(screenshot, (x1,y1), (x2,y2), (255,255,0), 1)
        cv2.imwrite("TT.png", screenshot)

        yan_question_QE_hw=yan_question_QE.shape[:2]
        x1 =  max_loc[0]-50
        x2 =  max_loc[0]-15
        y1 =  0
        y2 =  max_loc[1] + 20

        yan_question_QLT = yan_question_QE[y1:y2,x1:x2]
        yan_question_QLT_x1 = yan_question_QE_x1 +max_loc[0]-50
        yan_question_QLT_y1 = yan_question_QE_y1
        # yan_question_QLT_x1 = 
        # cv2.imwrite("yan_question_QLT.png", yan_question_QLT)
        ####
        if max_val>0.40:
            go_next_flag = 1
            print("yan_question_end HIGH",max_val)
        else: 
            go_next_flag = 0
            print("yan_question_R_bracket1 too low",max_val)
        
    if go_next_flag:
        image_to_find = cv2.imread("./pic/Yan/yan_question_LT1.png") #65
        # error: (-215:Assertion failed) _img.size().height <= _templ.size().height && _img.size().width <= _templ.size().width in function 'cv::matchTemplate'
        # if yan_question_QLT image size <= image_to_find image size go_next_flag=0
        if yan_question_QLT.shape[:2][0] <= image_to_find.shape[:2][0] or yan_question_QLT.shape[:2][1] <= image_to_find.shape[:2][1]:
            go_next_flag = 0
    if go_next_flag:
        image_to_find = cv2.imread("./pic/Yan/yan_question_LT1.png") #65
        result = cv2.matchTemplate(yan_question_QLT, image_to_find, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        # print(max_val)
        x1 =  max_loc[0]
        x2 =  max_loc[0]+13
        y1 =  max_loc[1]
        y2 =  max_loc[1]+24
        yan_question_QLT = cv2.rectangle(yan_question_QLT, (x1,y1), (x2,y2), (255,128,0), 1)

        x1 =  yan_question_QLT_x1 + max_loc[0]
        x2 =  yan_question_QLT_x1 + max_loc[0]+13
        y1 =  yan_question_QLT_y1 + max_loc[1]
        y2 =  yan_question_QLT_y1 + max_loc[1]+24

        screenshot = cv2.rectangle(screenshot, (x1,y1), (x2,y2), (0,128,0), 2)
        cv2.imwrite("TT.png", screenshot)

        cv2.imwrite("yan_question_QLT.png", yan_question_QLT)
        ####
        ####
        if max_val>0.10:
            go_next_flag = 1
            print("yan_question_end HIGH",max_val)
        else: 
            go_next_flag = 0
            print("yan_question_LT too low",max_val)
            time.sleep(1)
        
    if go_next_flag:
        yan_question_Q1 = screenshotPIL
        yan_question_QQ = yan_question_Q1
        # yan_question_QQ = cv2.cvtColor(numpy.array(yan_question_Q1), cv2.COLOR_RGB2BGR)
        
        xx1 =  x1 +5
        xx2 =  x1 +200
        yy1 =  y1+4
        yy2 =  y1+50
        yan_question_Q3 = yan_question_QQ[yy1:yy2,xx1:xx2]
        cv2.imwrite("yan_question_QQ3.png", yan_question_Q3)
        if 1:
            Question_confignow = ("-l chi_tra --oem 1 --psm 6")
            text3 = pytesseract.image_to_string(yan_question_Q3,config=Question_confignow)
            text3 = text_error_handle(text3)
            print("psm6= ",text3)
            Question_confignow = ("-l chi_tra --oem 1 --psm 7")
            text1 = pytesseract.image_to_string(yan_question_Q3,config=Question_confignow)
            print("psm7= ",text1)
        x1 =  x1
        x2 =  x1 +200
        y1 =  y1+4
        y2 =  y1+20
        yan_question_Q1 = yan_question_QQ[y1:y2,x1:x2]
        
        y1 =  y2 -1
        y2 =  y1+20
        yan_question_Q2 = yan_question_QQ[y1:y2,x1:x2]


        yan_question_QQ = cv2.hconcat([yan_question_Q1, yan_question_Q2])
        cv2.imwrite("yan_question_QQ.png", yan_question_QQ)
        # Question.save("Question_im1.png")

        # Question = get_screenshot(region=(Question_Loc_x,Question_Loc_y,Question_Loc_w,Question_Loc_l))
        if 1:
            Question_confignow = ("-l chi_tra --oem 1 --psm 6")
            text1 = pytesseract.image_to_string(yan_question_Q1,config=Question_confignow)
            text2 = pytesseract.image_to_string(yan_question_Q2,config=Question_confignow)
            print("psm6= ",text1+text2)
        Question_confignow = ("-l chi_tra --oem 1 --psm 7")
        text1 = pytesseract.image_to_string(yan_question_Q1,config=Question_confignow)
        text2 = pytesseract.image_to_string(yan_question_Q2,config=Question_confignow)
        print("psm7= ",text1+text2)

        text = text1 + text2
        text = text_error_handle(text)
        # print(text)
        # text=text
        text=text3

        pos = text.find("[")
        if pos < 8:
            text = text[pos+1:]

        pos = text.find("]")
        if pos < 8:
            numbers = re.findall(r'\d+', text)
            if numbers:
                numbers = list(numbers)
            if len(numbers)>0:
                Question_numbers = numbers[0]
            # print("Q text",text)
            text = text[pos+1:]
            text=text.replace("]","")
        Question_text = text
        print("find Q is ",Question_text)
        if Question_numbers != OLDQuestion_numbers:
            OLDQuestion_numbers = Question_numbers
            go_next_flag = 0
        elif Question_numbers == OLDQuestion_numbers:
            go_next_flag = 1
    #     screenshot = get_screenshot()
        # screenshot = screenshot.crop((0, 0, (screenshot.width), 700))
    #     screenshot = cv2.cvtColor(numpy.array(screenshot), cv2.COLOR_RGB2BGR)
        
    #     image_to_find = cv2.imread("./pic/Yan/RT20230129210406.png")
    #     result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
    #     min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    #     RT_loc = max_loc[0]+10
    #     image_to_find = cv2.imread("./pic/Yan/yan_question_LT.png") #80
    #     result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
    #     min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    #     if max_val>0.75:
            
    #         yan_question_LT_match_val = max_val
    #         yan_question_LT_match_loc = max_loc
    #         YAN_match_loc_x = yan_question_LT_match_loc[0]
    #         YAN_match_loc_y = yan_question_LT_match_loc[1]
    #         Question_Loc_x = YAN_match_loc_x+0
    #         Question_Loc_y = YAN_match_loc_y+5
    #         Question_Loc_w = RT_loc - YAN_match_loc_x +5
    #         Question_Loc_l = 20
    #         go_next_flag = 1
    #     else:
    #         go_next_flag = 0
    #         print("NO Q retry")
    # if go_next_flag:
    #     Question = get_screenshot(region=(Question_Loc_x,Question_Loc_y,Question_Loc_w,Question_Loc_l))
    #     Question.save("Question_im1.png")
    #     Question_confignow = ("-l chi_tra --oem 1 --psm 6")
    #     text1 = pytesseract.image_to_string(Question,config=Question_confignow)

    #     Question = get_screenshot(region=(Question_Loc_x,Question_Loc_y+20,Question_Loc_w,Question_Loc_l))
    #     Question.save("Question_im2.png")
    #     Question_confignow = ("-l chi_tra --oem 1 --psm 6")
    #     text2 = pytesseract.image_to_string(Question,config=Question_confignow)
    #     text = text1 + text2
    #     text=text.replace("\n",")
    #     text=text.replace(" ",")
    #     text=text.replace(")","]")
    #     text=text.replace("】","]")
    #     text=text.replace("}","]")

    #     text=text.replace("(","[")
    #     text=text.replace("【","[")
    #     text=text.replace("{","[")
    #     text = text_error_handle(text)
    #     pos = text.find("[")
    #     if pos < 8:
    #         text = text[pos+1:]

    #     pos = text.find("]")
    #     if pos < 8:
    #         numbers = re.findall(r'\d+', text)
    #         if numbers:
    #             numbers = list(numbers)
    #         if len(numbers)>0:
    #             Question_numbers = numbers[0]
    #         # print("Q text",text)
    #         text = text[pos+1:]
    #         text=text.replace("]",")
    #     Question_text = text

    #     # print("YYS Q text")
    #     # print(Question_text)


    # if go_next_flag:
    #     #get A    
    #     go_next_flag = 0


    #     image_to_find = cv2.imread("./pic/Yan/YAN24.png")
    #     result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
    #     min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    #     YAN24_match_val = max_val
    #     YAN24_match_loc = max_loc
    #     YAN24_im = get_screenshot(region=(max_loc[0],max_loc[1],100,100))
    #     YAN24_im.save("YAN24_im.png")
    #     if YAN24_match_val>0.8:
    #         go_next_flag = 0
    #     image_to_find = cv2.imread("./pic/Yan/YAN13.png")
    #     result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
    #     min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    #     YAN13_match_val = max_val
    #     YAN13_match_loc = max_loc
    #     YAN13_im = get_screenshot(region=(max_loc[0],max_loc[1],100,100))
    #     YAN13_im.save("YAN13_im.png")
    #     if YAN13_match_val>0.8 and YAN24_match_val>0.8:
    #         print("13 24 high enough",YAN13_match_val,YAN24_match_val)
    #         go_next_flag = 1
    #     elif (YAN13_match_val>0.75 or YAN24_match_val>0.75) and (YAN13_match_val>0.6 and YAN24_match_val>0.6):
    #         print("13 24 one very low try try see",YAN13_match_val,YAN24_match_val)
    #         go_next_flag = 1
    #     else:
    #         print("13 24 too low",YAN13_match_val,YAN24_match_val)
    #         pyautogui.moveTo(max_loc[0],max_loc[1])
    #         go_next_flag = 0
    
    ##########################################
    if go_next_flag:
        go_next_flag = 0
        # print(Question_text_lcut)
        sent2sim= []
        #get excel
        wb = load_workbook("./pic/FM/yys.xlsx")
        sheet = wb['Sheet1']
        my_list = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row+1)]
        EXCEL_ANS_list = [sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row+1)]
        #Question_text 分词
        
        # Question_text_lcut = jieba.lcut(Question_text)
        Question_text_lcut = set(list(str(Question_text)))

        if Question_text_lcut:
            go_next_flag = 1
        if go_next_flag:
            for sent2 in my_list:
                # EXCEL 分词
                # sent2_cut = jieba.lcut(sent2)
                set2 = set(list(str(sent2)))

                # 计算相似程度
                similarity = len(Question_text_lcut & set2) / len(Question_text_lcut | set2)
                # similarity = difflib.SequenceMatcher(None, Question_text_lcut, sent2_cut).quick_ratio()
                #add SIM in to list sent2sim
                sent2sim.append(similarity)
            # print("####2")
        similarity_max_value = 0
        if go_next_flag:
            FIND_ANS_list =[]
            similarity_max_value = max(sent2sim)
            if similarity_max_value < 0.1:
                go_next_flag=0
        if go_next_flag:
            # print("similarity_max_value",similarity_max_value)
            # # this is find index of MAX similarity sent2sim.index(similarity_max_value)
            similarity_max_index = sent2sim.index(similarity_max_value)

            for simii in range(len(sent2sim)):
                if similarity_max_value == sent2sim[simii]:
                    # sent2_cut = jieba.lcut(my_list[simii])
                    # print("index:",simii)
                    # print("my_list:",my_list[simii])
                    # print("sent2_cut:",sent2_cut)
                    # print("EXCEL_A:",EXCEL_ANS_list[simii])
                    print("EXCEL line[",simii,"],Q:[",my_list[simii],"],A:[",EXCEL_ANS_list[simii],"]")

                    FIND_ANS_list.append(EXCEL_ANS_list[simii])
            # print("##A##")
    ##########################################

    if go_next_flag:
        #### A1 
        
        ANS_confignow = ("-l chi_tra --oem 1 --psm 7")
        ANS_match_loc_w = 0
        print("len ans",len(FIND_ANS_list))
        if len(FIND_ANS_list) == 1:
            if isinstance(FIND_ANS_list[0],int):
                print("ans only number")
                ANS_confignow = ("-l eng --oem 1 --psm 7")
                if FIND_ANS_list[0] <10 :
                    print("ans number <10 set w is small")
                    ANS_match_loc_w = 10
            else:
                ANS_match_loc_w = len(FIND_ANS_list[0]) *13

                
        elif len(FIND_ANS_list) >= 2:
            FIND_ANS_len_max = 0
            for FIND_ANS_item in FIND_ANS_list:
                
                if isinstance(FIND_ANS_list[0],int):
                    ANS_match_loc_w = 20
                else:
                    if FIND_ANS_len_max<= len(FIND_ANS_item):
                        FIND_ANS_len_max = len(FIND_ANS_item)
            # if FIND_ANS_len_max<= 6:
            #     if isinstance(FIND_ANS_list[0],int):

            #     ANS_match_loc_w = len(FIND_ANS_list[0]) *15
        ################################################################
        
    if go_next_flag:
        x1 =  yan_question_QLT_x1 + max_loc[0]
        y1 =  yan_question_QLT_y1 + max_loc[1]
        yan_question_LT_x1 = x1-10
        yan_question_LT_x2 = x1+210
        yan_question_LT_y1 = y1

        yan_question_LT_y2 = y1+45
        yan_question_LT_y3 = y1+130
        screenshot = cv2.rectangle(screenshot, (yan_question_LT_x1,yan_question_LT_y1), (yan_question_LT_x2,yan_question_LT_y2), (255,255,0), 2)
        screenshot = cv2.rectangle(screenshot, (yan_question_LT_x1,yan_question_LT_y2), (yan_question_LT_x2,yan_question_LT_y3), (255,0,255), 2)

        ##find A

        # cv2.imwrite("TT.png", screenshot)
        A1_img=cv2.imread("./pic/Yan/A1.png")
        A2_img=cv2.imread("./pic/Yan/A2.png")
        A3_img=cv2.imread("./pic/Yan/A3.png")
        A4_img=cv2.imread("./pic/Yan/A4.png")
        # YYS_state = 3
        if YYS_state == 1:
            A1_img=cv2.imread("./pic/Yan/A1A.png")
            print("A1A")
        elif YYS_state == 2:
            A2_img=cv2.imread("./pic/Yan/A2A.png")
            print("A2A")
        elif YYS_state == 3:
            A3_img=cv2.imread("./pic/Yan/A3A.png")
            print("A3A")
        elif YYS_state == 4:
            A4_img=cv2.imread("./pic/Yan/A4A.png")
            print("A4A")
        #### FIND A13  :  screenshot = cv2.rectangle X4
        #### FIND A13
        ##find A3
        yan_question_Q = screenshotPIL
        # yan_question_Q = yan_question_Q.crop((yan_question_LT_x1, yan_question_LT_y2, yan_question_LT_x1+70, yan_question_LT_y3))
        
    # if go_next_flag:
        # error: (-215:Assertion failed) _img.size().height <= _templ.size().height && _img.size().width <= _templ.size().width in function 'cv::matchTemplate'
        # if yan_question_QLT image size <= image_to_find image size go_next_flag=0
        # if yan_question_Q.shape[:2][0] <= 20 or yan_question_Q.shape[:2][1] <= 20:
        #     go_next_flag = 0
            
    if go_next_flag:
        image_to_find = A3_img
        
        # yan_question_Q = cv2.cvtColor(numpy.array(yan_question_Q), cv2.COLOR_RGB2BGR)
        #show yan_question_Q in a new window
        # cv2.imshow("yan_question_Q",yan_question_Q)
        # cv2.waitKey(0)

        result = cv2.matchTemplate(yan_question_Q, image_to_find, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        print("A3_img",max_val)
        A3x1 = yan_question_LT_x1+max_loc[0]
        A3x2 = yan_question_LT_x1+max_loc[0]+60
        A3y1 = yan_question_LT_y2+max_loc[1]
        A3y2 = yan_question_LT_y2+max_loc[1]+16
        screenshot = cv2.rectangle(screenshot, (A3x1,A3y1), (A3x2,A3y2), (0,255,0), 2)

        ##find A1
        yan_question_Q = screenshotPIL
        yan_question_Q = Image.fromarray(yan_question_Q)
        yan_question_Q = yan_question_Q.crop((yan_question_LT_x1, yan_question_LT_y2, yan_question_LT_x1+70, A3y1))
        image_to_find = A1_img
        
        yan_question_Q = cv2.cvtColor(numpy.array(yan_question_Q), cv2.COLOR_RGB2BGR)
        result = cv2.matchTemplate(yan_question_Q, image_to_find, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        print("A1_img",max_val)
        A1x1 = yan_question_LT_x1+max_loc[0]
        A1x2 = yan_question_LT_x1+max_loc[0]+60
        A1y1 = yan_question_LT_y2+max_loc[1]
        A1y2 = yan_question_LT_y2+max_loc[1]+10
        screenshot = cv2.rectangle(screenshot, (A1x1,A1y1), (A1x2,A1y2), (0,0,255), 1)



        #### FIND A24
        ##find A4
        yan_question_Q = screenshotPIL
        yan_question_Q = Image.fromarray(yan_question_Q)
        yan_question_Q = yan_question_Q.crop((yan_question_LT_x1+70, yan_question_LT_y2, yan_question_LT_x2, yan_question_LT_y3))
        image_to_find = A4_img
        
        yan_question_Q = cv2.cvtColor(numpy.array(yan_question_Q), cv2.COLOR_RGB2BGR)
        result = cv2.matchTemplate(yan_question_Q, image_to_find, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        print("A4_img",max_val)
        A4x1 = yan_question_LT_x1+70+max_loc[0]
        # A4x2 = yan_question_LT_x1+70+max_loc[0]+60
        A4x2 = yan_question_LT_x2
        A4y1 = yan_question_LT_y2+max_loc[1]
        A4y2 = yan_question_LT_y2+max_loc[1]+16
        screenshot = cv2.rectangle(screenshot, (A4x1,A4y1), (A4x2,A4y2), (128,128,0), 1)

        ##find A2
        yan_question_Q = screenshotPIL
        yan_question_Q = Image.fromarray(yan_question_Q)
        yan_question_Q = yan_question_Q.crop((yan_question_LT_x1+70, yan_question_LT_y2, yan_question_LT_x2, yan_question_LT_y3))
        image_to_find = A2_img
        
        yan_question_Q = cv2.cvtColor(numpy.array(yan_question_Q), cv2.COLOR_RGB2BGR)
        result = cv2.matchTemplate(yan_question_Q, image_to_find, cv2.TM_CCOEFF_NORMED)
        min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
        print("A2_img",max_val)
        A2x1 = yan_question_LT_x1+70+max_loc[0]
        # A2x2 = yan_question_LT_x1+70+max_loc[0]+60
        A2x2 = yan_question_LT_x2
        A2y1 = yan_question_LT_y2+max_loc[1]
        A2y2 = yan_question_LT_y2+max_loc[1]+16
        screenshot = cv2.rectangle(screenshot, (A2x1,A2y1), (A2x2,A2y2), (0,128,128), 1)

        A1x1 = A1x1    +10
        # A1x2 = A1x2    
        A1x2 = A2x1    -5
        A1y1 = A1y1    -3 +1
        A1y2 = A1y2    +3 +1

        A2x1 = A2x1    +16 
        A2x2 = A2x2    
        A2y1 = A2y1    -0
        A2y2 = A2y2    -0

        A3x1 = A3x1    +16
        # A3x2 = A3x2    
        A3x2 = A4x1    -5
        A3y1 = A3y1    +1
        A3y2 = A3y2    +1
    
        A4x1 = A4x1    +16 +2
        A4x2 = A4x2    
        A4y1 = A4y1    
        A4y2 = A4y2    
    
        if ANS_match_loc_w:
            A1x2 = A1x1 + ANS_match_loc_w
            A2x2 = A2x1 + ANS_match_loc_w
            A3x2 = A3x1 + ANS_match_loc_w
            A4x2 = A4x1 + ANS_match_loc_w
        if BaseOnStateMoveAns:
            A1y1 = A1y1 + BaseOnStateMoveMatrix[YYS_state-1][1-1]
            A1y2 = A1y2 + BaseOnStateMoveMatrix[YYS_state-1][1-1]
            A2y1 = A2y1 + BaseOnStateMoveMatrix[YYS_state-1][2-1]
            A2y2 = A2y2 + BaseOnStateMoveMatrix[YYS_state-1][2-1]
            A3y1 = A3y1 + BaseOnStateMoveMatrix[YYS_state-1][3-1]
            A3y2 = A3y2 + BaseOnStateMoveMatrix[YYS_state-1][3-1]
            A4y1 = A4y1 + BaseOnStateMoveMatrix[YYS_state-1][4-1]
            A4y2 = A4y2 + BaseOnStateMoveMatrix[YYS_state-1][4-1]

        

        ANS1 = screenshot[A1y1:A1y2,A1x1:A1x2]
        ANS2 = screenshot[A2y1:A2y2,A2x1:A2x2]
        ANS3 = screenshot[A3y1:A3y2,A3x1:A3x2]
        ANS4 = screenshot[A4y1:A4y2,A4x1:A4x2]

        # screenshot = cv2.rectangle(screenshot, (A1x1,A1y1), (A1x2,A1y2), (0,0,255), 1)
        # screenshot = cv2.rectangle(screenshot, (A3x1,A3y1), (A3x2,A3y2), (0,255,0), 1)
        # screenshot = cv2.rectangle(screenshot, (A2x1,A2y1), (A2x2,A2y2), (0,128,128), 1)
        # screenshot = cv2.rectangle(screenshot, (A4x1,A4y1), (A4x2,A4y2), (128,128,0), 1)

        ANS_img = cv2.hconcat([ANS1, ANS2,ANS3,ANS4])
        cv2.imwrite("ANS_img.png", ANS_img)

        cv2.imwrite("TT.png", screenshot)
        

        #################################################################
        # ANS1 = get_screenshot(region=(ANS1_match_loc_x,ANS1_match_loc_y,ANS1_match_loc_w,ANS1_match_loc_l))
        # ANS1.save("ANS1_im.png")
        text = pytesseract.image_to_string(ANS1,config=ANS_confignow)
        text=text.replace("\n","")
        text=text.replace(" ","")
        text=text.replace(",","")
        text = text_error_handle(text)
        ANS1_text = text
        # print("text1")
        # print(text)

        # ANS3 = get_screenshot(region=(ANS3_match_loc_x,ANS3_match_loc_y,ANS3_match_loc_w,ANS3_match_loc_l))
        # ANS3.save("ANS3_im.png")
        text = pytesseract.image_to_string(ANS3,config=ANS_confignow)
        text=text.replace("\n","")
        text=text.replace(" ","")
        text=text.replace(",","")
        text = text_error_handle(text)
        ANS3_text = text
        # print("text3")
        # print(text)

        # ANS2 = get_screenshot(region=(ANS2_match_loc_x,ANS2_match_loc_y,ANS2_match_loc_w,ANS2_match_loc_l))
        # ANS2.save("ANS2_im.png")
        text = pytesseract.image_to_string(ANS2,config=ANS_confignow)
        text=text.replace("\n","")
        text=text.replace(" ","")
        text=text.replace(",","")
        text = text_error_handle(text)
        ANS2_text = text
        # print("text2")
        # print(text)
        

        # ANS4 = get_screenshot(region=(ANS4_match_loc_x,ANS4_match_loc_y,ANS4_match_loc_w,ANS4_match_loc_l))
        # ANS4.save("ANS4_im.png")
        text = pytesseract.image_to_string(ANS4,config=ANS_confignow)
        text=text.replace("\n","")
        text=text.replace(" ","")
        text=text.replace(",","")
        text = text_error_handle(text)
        ANS4_text = text
        # print("text4")
        # print(text)


        #################################################################
        # ANS1.save("ANS1_im.png")
        # ANS3.save("ANS3_im.png")
        # ANS2.save("ANS2_im.png")
        # ANS4.save("ANS4_im.png")
        YYS_ANS_list = [ANS1_text,ANS2_text,ANS3_text,ANS4_text]
        # print("YYS A LIST")
        print("YYS A LIST",YYS_ANS_list)
        go_next_flag = 1
        ##########################################
        if go_next_flag:
            for FIND_ANS_item in FIND_ANS_list:
                FIND_ANS_item_cut = set(list(str(FIND_ANS_item)))
                if isinstance(FIND_ANS_list[0],int):
                    print("ans only number")
                # else:
                #     for YYS_ANS_item in YYS_ANS_list:
                #         if len(FIND_ANS_item)> len(YYS_ANS_item):
                #             YYS_ANS_list[YYS_ANS_list.index(YYS_ANS_item)] = YYS_ANS_item +"xo"
            
            # print("#")
            # print("#")
            similarity_max = 0
            th_hold = 0.3
            YYS_possible_ANS_list = []
            if 0:#isinstance(FIND_ANS_list[0], (int,float)):
                # print("ANS NO CHINESE")
                for FIND_ANS_item in FIND_ANS_list:
                    for YYS_ANS_item in YYS_ANS_list:
                        if FIND_ANS_item == YYS_ANS_item:
                            print("maybe index [",YYS_ANS_list.index(YYS_ANS_item)+1,"],ans is [",FIND_ANS_item,"]")
            else:
                for FIND_ANS_item in FIND_ANS_list:
                    # FIND_ANS_item_cut = jieba.lcut(FIND_ANS_item)
                    FIND_ANS_item_cut = set(list(str(FIND_ANS_item)))
                    for YYS_ANS_item in YYS_ANS_list:
                        # if len(FIND_ANS_item)> len(YYS_ANS_item):
                        #     # YYS_ANS_list[YYS_ANS_list.index(YYS_ANS_item)] = YYS_ANS_item +"xo"
                        #     # YYS_ANS_item = YYS_ANS_item +"xo"
                        #     YYS_ANS_item = YYS_ANS_item
                        # YYS_ANS_item_cut = jieba.lcut(YYS_ANS_item)
                        YYS_ANS_item_cut = set(list(str(YYS_ANS_item)))
                        similarity = len(FIND_ANS_item_cut & YYS_ANS_item_cut) / len(FIND_ANS_item_cut | YYS_ANS_item_cut)
                        # similarity = difflib.SequenceMatcher(None, FIND_ANS_item_cut, YYS_ANS_item_cut).quick_ratio()
                        # print(FIND_ANS_item_cut,";",YYS_ANS_item_cut,"S:",similarity)
                        if similarity_max < similarity:
                            similarity_max = similarity
                            YYS_possible_ANS_list = []
                        if similarity_max == similarity:
                            YYS_possible_ANS_list.append(YYS_ANS_item)
                print("possible ans ",YYS_possible_ANS_list)
                # for FIND_ANS_item in FIND_ANS_list:
                #     # FIND_ANS_item_cut = jieba.lcut(FIND_ANS_item)
                #     FIND_ANS_item_cut = set(list(str(FIND_ANS_item)))
                #     for YYS_ANS_item in YYS_ANS_list:
                #         # YYS_ANS_item_cut = jieba.lcut(YYS_ANS_item)
                #         YYS_ANS_item_cut = set(list(str(YYS_ANS_item)))
                #         # similarity = difflib.SequenceMatcher(None, FIND_ANS_item_cut, YYS_ANS_item_cut).quick_ratio()
                #         similarity = len(FIND_ANS_item_cut & YYS_ANS_item_cut) / len(FIND_ANS_item_cut | YYS_ANS_item_cut)
                #         if (similarity_max == similarity) and (similarity_max > (th_hold-0.1)):
                #             print(similarity_max,"maybe index [",YYS_ANS_list.index(YYS_ANS_item)+1,"],ans is [",FIND_ANS_item,"]")
                #             YYS_possible_ANS_list.append(YYS_ANS_item)
                if len(YYS_possible_ANS_list) == 1:
                    YYS_ANS_item=YYS_possible_ANS_list[0]
                    print("choose only one",YYS_ANS_list.index(YYS_ANS_item)+1)
                    movebystate(YYS_ANS_list.index(YYS_ANS_item)+1)
                elif len(YYS_possible_ANS_list) == 2:
                    print("choose first one")
                    movebystate(YYS_ANS_list.index(YYS_ANS_item)+1)
                    timenow = time.strftime('%m%d_%H%M%S',time.localtime(time.time()))
                    cv2.imwrite("yan_BAD2_ans_"+str(timenow)+".png", screenshot)
                elif len(YYS_possible_ANS_list) == 3:# choose left one
                    print("choose left one")
                    newYYS_ANS_list = YYS_ANS_list
                    # for YYS_possible_ANS_item in YYS_possible_ANS_list:
                        # newYYS_ANS_list.remove(YYS_possible_ANS_item)
                    YYS_ANS_item=newYYS_ANS_list[0]
                    movebystate(YYS_ANS_list.index(YYS_ANS_item)+1)
                    timenow = time.strftime('%m%d_%H%M%S',time.localtime(time.time()))
                    cv2.imwrite("yan_BAD3_ans_"+str(timenow)+".png", screenshot)
                else:
                    print("do not thing")
                    timenow = time.strftime('%m%d_%H%M%S',time.localtime(time.time()))
                    cv2.imwrite("yan_BAD4_ans_"+str(timenow)+".png", screenshot)
                if 0:
                    ####
                    for FIND_ANS_item in FIND_ANS_list:
                        # FIND_ANS_item_cut = jieba.lcut(FIND_ANS_item)
                        FIND_ANS_item_cut = set(list(str(FIND_ANS_item)))
                        for YYS_ANS_item in YYS_ANS_list:
                            # YYS_ANS_item_cut = jieba.lcut(YYS_ANS_item)
                            YYS_ANS_item_cut = set(list(str(YYS_ANS_item)))
                            # similarity = difflib.SequenceMatcher(None, FIND_ANS_item_cut, YYS_ANS_item_cut).quick_ratio()
                            similarity = len(FIND_ANS_item_cut & YYS_ANS_item_cut) / len(FIND_ANS_item_cut | YYS_ANS_item_cut)
                            if (similarity_max == similarity) and (similarity_max > (th_hold-0.1)):
                                print(similarity_max,"maybe index [",YYS_ANS_list.index(YYS_ANS_item)+1,"],ans is [",FIND_ANS_item,"]")
                                movebystate(YYS_ANS_list.index(YYS_ANS_item)+1)
            if similarity_max < th_hold:
                Question_numbersOLD = 0
            timenow = time.strftime('%m%d_%H%M%S',time.localtime(time.time()))
            cv2.imwrite("yan_ans_"+str(timenow)+".png", screenshot)
            # print("EXCEL count",len(FIND_ANS_list)," ANS",FIND_ANS_list)
            # print(YYS_ANS_list)
            
    time.sleep(0.1)








ADicon = cv2.imread("./pic/Yan/android_icon.png")
image_to_find = cv2.imread("./pc/Yan/./pic/Yan/yan_question_end.png") #70
image_to_find = cv2.imread("./pc/Yan/./pic/Yan/yan_question_R_bracket1.png") #40
image_to_find = cv2.imread("./pc/Yan/./pic/Yan/yan_question_LT.png") #65
A1_img=cv2.imread("./pc/Yan/./pic/Yan/A1.png")
A2_img=cv2.imread("./pc/Yan/./pic/Yan/A2.png")
A3_img=cv2.imread("./pc/Yan/./pic/Yan/A3.png")
A4_img=cv2.imread("./pc/Yan/./pic/Yan/A4.png")
A1_img=cv2.imread("./pc/Yan/./pic/Yan/A1A.png")
A2_img=cv2.imread("./pc/Yan/./pic/Yan/A2A.png")
A3_img=cv2.imread("./pc/Yan/./pic/Yan/A3A.png")
A4_img=cv2.imread("./pc/Yan/./pic/Yan/A4A.png")
YanHui_XYLL_N = cv2.imread("./pc/Yan/./pic/Yan/1666177652310.png")
YanHui_LiaoXinXi = cv2.imread("./pc/Yan/./pic/Yan/new_YYL_UI_LiaoXinXi.png")
YanHui_XYLL_N = cv2.imread("./pc/Yan/./pic/Yan/1673188292135.png")
YanHui_XYLL_N = cv2.imread("./pc/Yan/./pic/Yan/new_YYL_UI_LiaoXinXi.png")
YanHui_XYLL_N = cv2.imread("./pc/Yan/./pic/Yan/new_YYL_UI_LiaoXinXi.png")
YanHui_XYLL_Y = cv2.imread("./pc/Yan/./pic/Yan/1673188475535.png")
image_to_find = cv2.imread("./pc/Yan/./pic/Yan/yan_question_end.png") #70
image_to_find = cv2.imread("./pc/Yan/./pic/Yan/yan_question_R_bracket1.png") #40
image_to_find = cv2.imread("./pc/Yan/./pic/Yan/yan_question_R_bracket1.png") #65
image_to_find = cv2.imread("./pc/Yan/./pic/Yan/yan_question_LT.png") #65
image_to_find = cv2.imread("./pc/Yan/./pic/Yan/yan_question_LT.png") #65
A1_img=cv2.imread("./pc/Yan/./pic/Yan/A1.png")
A2_img=cv2.imread("./pc/Yan/./pic/Yan/A2.png")
A3_img=cv2.imread("./pc/Yan/./pic/Yan/A3.png")
A4_img=cv2.imread("./pc/Yan/./pic/Yan/A4.png")
A1_img=cv2.imread("./pc/Yan/./pic/Yan/A1A.png")
A2_img=cv2.imread("./pc/Yan/./pic/Yan/A2A.png")
A3_img=cv2.imread("./pc/Yan/./pic/Yan/A3A.png")
A4_img=cv2.imread("./pc/Yan/./pic/Yan/A4A.png")
