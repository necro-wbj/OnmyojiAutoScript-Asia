campare_tool = 2 #1 jieba #2 math 
Do_Click_Flag = 1
GO_next_step = 1
click_mail = 1

import adbutils
import pyautogui
import pytesseract
from openpyxl import load_workbook
import jieba
import difflib
import cv2
import numpy
import time
import numpy as np

import BASE

MAX_similarity = 123
YYS_Q_text = "123"
ORC_YYS_ANS_list = "456"

YYS_XLSX_ADDR = "D:\PROJ\PY_android\pic\FM\yys.xlsx"

def text_error_handle(text):
    text = text.replace("御魂 狠 ", "禦魂 猙 ")
    text = text.replace("蕙", "薰")
    text = text.replace("蔥", "薰")
    text = text.replace("自", "白")
    text = text.replace("\n","")
    text = text.replace("|","")
    text = text.replace(" ","")
    text = text.replace("「", " ")
    text = text.replace("」", " ")
    text = text.replace("'", " ")
    text = text.replace("《", " ")
    text = text.replace("》", " ")
    return text

#jieba
# jieba.load_userdict("userdict.txt")
#pytesseract
print(time.ctime(time.time()))
confignow = ("-l chi_tra --oem 1 --psm 7")
FM_FengMoMiXin = "./pic/FM/FM_FengMoMiXin.png"
YYS_QuDeJiangLi = "./pic/base/YYS_QuDeJiangLi.png"



if 1:
    devices = adbutils.adb.device_list()
    for device in devices:
        FM_MAIL = "./pic/FM/FM4_MAIL.png"
        x,y = BASE.exists(device,FM_MAIL,0.60)
        while x or y:
            BASE.click_point(device,x,y)
            time.sleep(1)
            x1,y1 = BASE.exists(device,FM_FengMoMiXin,0.75)
            # click FM4_MAIL pass and find FM_FengMoMiXin
            if x1 or y1:
                BASE.click_point(device,x,y)


                # SS = "./pic/FM/SS.png"
                # SS = cv2.imread(SS)
                # screenshot = SS
                # SS = cv2.cvtColor(SS, cv2.COLOR_RGB2BGR) already BGR no need this
                #show SS
                screen = device.screenshot()
                screen = cv2.cvtColor(np.array(screen), cv2.COLOR_RGB2BGR)
                screenshot = screen


                image_to_find = cv2.imread(FM_FengMoMiXin)
                result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
                min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
                FMMX_Loc_x = max_loc[0]-4
                FMMX_Loc_y = max_loc[1]+3

                Question1_Loc_x = FMMX_Loc_x
                Question1_Loc_y = FMMX_Loc_y + 75
                Question1_Loc_w = 340
                Question1_Loc_l = 23
                #draw a box of Question_Loc
                # cv2.rectangle(screenshot, (Question1_Loc_x, Question1_Loc_y), (Question1_Loc_x+Question1_Loc_w, Question1_Loc_y+Question1_Loc_l), (0, 0, 255), 1)
                #cut screenshot to Question1_Loc and show
                Question1_Loc = screenshot[Question1_Loc_y:Question1_Loc_y+Question1_Loc_l, Question1_Loc_x:Question1_Loc_x+Question1_Loc_w]
                # cv2.imshow("Question1_Loc",Question1_Loc)
                # cv2.waitKey(0)

                Question2_Loc_x = FMMX_Loc_x
                Question2_Loc_y = FMMX_Loc_y + 97
                Question2_Loc_w = 340
                Question2_Loc_l = 25
                #draw a box of Question2_Loc with green
                # cv2.rectangle(screenshot, (Question2_Loc_x, Question2_Loc_y), (Question2_Loc_x+Question2_Loc_w, Question2_Loc_y+Question2_Loc_l), (0, 255, 0), 1)
                #cut screenshot to Question2_Loc and show
                Question2_Loc = screenshot[Question2_Loc_y:Question2_Loc_y+Question2_Loc_l, Question2_Loc_x:Question2_Loc_x+Question2_Loc_w]
                # cv2.imshow("Question2_Loc",Question2_Loc)
                # cv2.waitKey(0)


                Ans_Loc_w = 290
                Ans_Loc_l = 35
                Ans1_Loc_x = FMMX_Loc_x + 30
                Ans1_Loc_y = FMMX_Loc_y + 155

                # Ans1 = pyautogui.screenshot(region=(Ans1_Loc_x,Ans1_Loc_y,Ans_Loc_w,Ans_Loc_l))
                Ans1 =screenshot[Ans1_Loc_y:Ans1_Loc_y+Ans_Loc_l, Ans1_Loc_x:Ans1_Loc_x+Ans_Loc_w]
                # cv2.imshow("Ans1",Ans1)
                # cv2.waitKey(0)

                Ans2_Loc_x = FMMX_Loc_x + 30
                Ans2_Loc_y = FMMX_Loc_y + 225
                # Ans2 = pyautogui.screenshot(region=(Ans2_Loc_x,Ans2_Loc_y,Ans_Loc_w,Ans_Loc_l))
                # Ans2.save("Ans2.png")
                Ans2 = screenshot[Ans2_Loc_y:Ans2_Loc_y+Ans_Loc_l, Ans2_Loc_x:Ans2_Loc_x+Ans_Loc_w]
                # cv2.imshow("Ans2",Ans2)
                # cv2.waitKey(0)

                Ans3_Loc_x = FMMX_Loc_x + 30
                Ans3_Loc_y = FMMX_Loc_y + 285

                # Ans3 = pyautogui.screenshot(region=(Ans3_Loc_x,Ans3_Loc_y,Ans_Loc_w,Ans_Loc_l))
                # Ans3.save("Ans3.png")
                Ans3 = screenshot[Ans3_Loc_y:Ans3_Loc_y+Ans_Loc_l, Ans3_Loc_x:Ans3_Loc_x+Ans_Loc_w]
                # cv2.imshow("Ans3",Ans3)
                # cv2.waitKey(0)



                confignow = ("-l chi_tra --oem 1 --psm 7")
                #Q to text
                
                text1 = pytesseract.image_to_string(Question1_Loc,config=confignow)
                text2 = pytesseract.image_to_string(Question2_Loc,config=confignow)
                text = text1 +text2
                text = text_error_handle(text)
                # print(text)

                YYS_Q_text = text
                # print("#### YYS Question text is ",YYS_Q_text)

                #find text in excel
                wb = load_workbook(YYS_XLSX_ADDR)
                sheet = wb["Sheet1"]
                excel_Q_list = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row+1)]
                excel_ANS_list = [sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row+1)]
                sent1_cut = list(str(YYS_Q_text))
                set1 = set(sent1_cut)
                MAXsimilarity = 0
                for sent2 in excel_Q_list:
                    similarity = 0
                    sent2_cut = list(str(sent2))
                    set2 = set(sent2_cut)
                    #calculate text similarity
                    similarity = difflib.SequenceMatcher(None, sent1_cut, sent2_cut).quick_ratio()
                    # find max sim
                    if similarity > MAXsimilarity:
                        MAXsimilarity = similarity
                correct_excel_Que_list = []
                correct_excel_ANS_list = []
                # save MAX sim to list

                for sent2index in range(len(excel_Q_list)):
                    similarity = 0
                    sent2_cut = list(str(excel_Q_list[sent2index]))
                    set2 = set(sent2_cut)
                    #calculate text similarity
                    similarity = difflib.SequenceMatcher(None, sent1_cut, sent2_cut).quick_ratio()
                    # find max sim
                    if similarity == MAXsimilarity:
                        correct_excel_Que_list.append(excel_Q_list[sent2index])
                        correct_excel_ANS_list.append(excel_ANS_list[sent2index])
                
                        print(sent2index,excel_Q_list[sent2index],excel_ANS_list[sent2index])

                text = pytesseract.image_to_string(Ans1,config=confignow)
                text1 = text_error_handle(text)
                # print(text)
                text = pytesseract.image_to_string(Ans2,config=confignow)
                text2 = text_error_handle(text)
                # print(text)
                text = pytesseract.image_to_string(Ans3,config=confignow)
                text3 = text_error_handle(text)
                # print(text)
                ORC_YYS_ANS_list = [text1,text2,text3]
                print("ORC YYS Qus: ",YYS_Q_text)
                print("ORC YYS Ans: ",ORC_YYS_ANS_list)


                if 1:
                    MAX_similarity=0
                    CORRECT_INDEX = 0
                    # print(correct_excel_ANS_list)
                    for correct_ANS_item in correct_excel_ANS_list:
                        for ORC_ANS_item in ORC_YYS_ANS_list:
                            sent1_cut = set(list(str(correct_ANS_item)))
                            sent2_cut = set(list(str(ORC_ANS_item)))
                            similarity = difflib.SequenceMatcher(None, sent1_cut, sent2_cut).quick_ratio()
                            if similarity > MAX_similarity:
                                MAX_similarity = similarity
                    for correct_ANS_item in correct_excel_ANS_list:
                        for ORC_ANS_item in ORC_YYS_ANS_list:
                            sent1_cut = set(list(str(correct_ANS_item)))
                            sent2_cut = set(list(str(ORC_ANS_item)))
                            similarity = difflib.SequenceMatcher(None, sent1_cut, sent2_cut).quick_ratio()
                            if similarity == MAX_similarity:
                                CORRECT_INDEX = ORC_YYS_ANS_list.index(ORC_ANS_item)
                                print("####sim= ",similarity,",YYS Ans is ",CORRECT_INDEX,ORC_ANS_item)
                            if MAX_similarity<0.5:
                                # save YYS_Q_text and ORC_YYS_ANS_list to "./pic/FM/error.txt" with UTF8
                                with open("./pic/FM/error.txt","a",encoding='utf-8') as f:
                                    MAX_similarity=MAX_similarity+87
                                    f.write("simLOW="+str(MAX_similarity)+"\n")
                                    f.write(str(YYS_Q_text)+str(ORC_YYS_ANS_list)+"\n")
                                    f.write("\n")
                
                    Ans1_click_Loc_x = 150 + FMMX_Loc_x + 30
                    Ans1_click_Loc_y = 15+FMMX_Loc_y + 155
                    Ans2_click_Loc_x = 150 + FMMX_Loc_x + 30
                    Ans2_click_Loc_y = 15+FMMX_Loc_y + 225
                    Ans3_click_Loc_x = 150 + FMMX_Loc_x + 30
                    Ans3_click_Loc_y = 15+FMMX_Loc_y + 285

                    #draw circle for Ans1_click_Loc_x in screenshot
                    # cv2.circle(screenshot,(Ans1_click_Loc_x,Ans1_click_Loc_y), 10, (0,0,255), -1)
                    # cv2.circle(screenshot,(Ans2_click_Loc_x,Ans2_click_Loc_y), 10, (0,0,255), -1)
                    # cv2.circle(screenshot,(Ans3_click_Loc_x,Ans3_click_Loc_y), 10, (0,0,255), -1)
                    Ans_click_Loc_x = [Ans1_click_Loc_x,Ans2_click_Loc_x,Ans3_click_Loc_x]
                    Ans_click_Loc_y = [Ans1_click_Loc_y,Ans2_click_Loc_y,Ans3_click_Loc_y]
                    # cv2.circle(screenshot,(Ans_click_Loc_x[CORRECT_INDEX],Ans_click_Loc_y[CORRECT_INDEX]), 10, (0,0,255), -1)
                    #click Ans
                    BASE.click_point(device,Ans_click_Loc_x[CORRECT_INDEX],Ans_click_Loc_y[CORRECT_INDEX])
                    time.sleep(1)
                    x2 = 0
                    y2 = 0
                    x2,y2 = BASE.wait_picture(device,YYS_QuDeJiangLi,th_hold=0.95,wait_time=10)
                    if x2 or y2:
                        x2 = x2+300
                        y2 = y2-50
                        BASE.click_point(device,x2,y2)
                    else:
                        # save YYS_Q_text and ORC_YYS_ANS_list to "./pic/FM/error.txt" with UTF8
                        with open("./pic/FM/error.txt","a",encoding='utf-8') as f:
                            
                            f.write("sim="+str(MAX_similarity)+"\n")
                            f.write(str(YYS_Q_text)+str(ORC_YYS_ANS_list)+"\n")
                            f.write("\n")


                # cv2.imshow("SS",SS)
                # cv2.waitKey(0)


            else:
                BASE.click_point(device,x,y)
            x,y = BASE.exists(device,FM_MAIL,th_hold = 0.6)





###############
if 0:
    SS = "./pic/FM/SS.png"
    SS = cv2.imread(SS)
    screenshot = SS
    # SS = cv2.cvtColor(SS, cv2.COLOR_RGB2BGR) already BGR no need this
    #show SS

    image_to_find = cv2.imread(FM_FengMoMiXin)
    result = cv2.matchTemplate(screenshot, image_to_find, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    FMMX_Loc_x = max_loc[0]-4
    FMMX_Loc_y = max_loc[1]+3

    Question1_Loc_x = FMMX_Loc_x
    Question1_Loc_y = FMMX_Loc_y + 75
    Question1_Loc_w = 340
    Question1_Loc_l = 23
    #draw a box of Question_Loc
    # cv2.rectangle(screenshot, (Question1_Loc_x, Question1_Loc_y), (Question1_Loc_x+Question1_Loc_w, Question1_Loc_y+Question1_Loc_l), (0, 0, 255), 1)
    #cut screenshot to Question1_Loc and show
    Question1_Loc = screenshot[Question1_Loc_y:Question1_Loc_y+Question1_Loc_l, Question1_Loc_x:Question1_Loc_x+Question1_Loc_w]
    # cv2.imshow("Question1_Loc",Question1_Loc)
    # cv2.waitKey(0)

    Question2_Loc_x = FMMX_Loc_x
    Question2_Loc_y = FMMX_Loc_y + 97
    Question2_Loc_w = 340
    Question2_Loc_l = 25
    #draw a box of Question2_Loc with green
    # cv2.rectangle(screenshot, (Question2_Loc_x, Question2_Loc_y), (Question2_Loc_x+Question2_Loc_w, Question2_Loc_y+Question2_Loc_l), (0, 255, 0), 1)
    #cut screenshot to Question2_Loc and show
    Question2_Loc = screenshot[Question2_Loc_y:Question2_Loc_y+Question2_Loc_l, Question2_Loc_x:Question2_Loc_x+Question2_Loc_w]
    # cv2.imshow("Question2_Loc",Question2_Loc)
    # cv2.waitKey(0)


    Ans_Loc_w = 290
    Ans_Loc_l = 35
    Ans1_Loc_x = FMMX_Loc_x + 30
    Ans1_Loc_y = FMMX_Loc_y + 155

    # Ans1 = pyautogui.screenshot(region=(Ans1_Loc_x,Ans1_Loc_y,Ans_Loc_w,Ans_Loc_l))
    Ans1 =screenshot[Ans1_Loc_y:Ans1_Loc_y+Ans_Loc_l, Ans1_Loc_x:Ans1_Loc_x+Ans_Loc_w]
    # cv2.imshow("Ans1",Ans1)
    # cv2.waitKey(0)

    Ans2_Loc_x = FMMX_Loc_x + 30
    Ans2_Loc_y = FMMX_Loc_y + 225
    # Ans2 = pyautogui.screenshot(region=(Ans2_Loc_x,Ans2_Loc_y,Ans_Loc_w,Ans_Loc_l))
    # Ans2.save("Ans2.png")
    Ans2 = screenshot[Ans2_Loc_y:Ans2_Loc_y+Ans_Loc_l, Ans2_Loc_x:Ans2_Loc_x+Ans_Loc_w]
    # cv2.imshow("Ans2",Ans2)
    # cv2.waitKey(0)

    Ans3_Loc_x = FMMX_Loc_x + 30
    Ans3_Loc_y = FMMX_Loc_y + 285

    # Ans3 = pyautogui.screenshot(region=(Ans3_Loc_x,Ans3_Loc_y,Ans_Loc_w,Ans_Loc_l))
    # Ans3.save("Ans3.png")
    Ans3 = screenshot[Ans3_Loc_y:Ans3_Loc_y+Ans_Loc_l, Ans3_Loc_x:Ans3_Loc_x+Ans_Loc_w]
    # cv2.imshow("Ans3",Ans3)
    # cv2.waitKey(0)



    confignow = ("-l chi_tra --oem 1 --psm 7")
    #Q to text
    
    text1 = pytesseract.image_to_string(Question1_Loc,config=confignow)
    text2 = pytesseract.image_to_string(Question2_Loc,config=confignow)
    text = text1 +text2
    text = text_error_handle(text)
    # print(text)

    YYS_Q_text = text
    print("#### YYS Question text is ",YYS_Q_text)

    #find text in excel
    wb = load_workbook(YYS_XLSX_ADDR)
    sheet = wb["Sheet1"]
    excel_Q_list = [sheet.cell(row=i, column=1).value for i in range(1, sheet.max_row+1)]
    excel_ANS_list = [sheet.cell(row=i, column=2).value for i in range(1, sheet.max_row+1)]
    sent1_cut = list(str(YYS_Q_text))
    set1 = set(sent1_cut)
    MAXsimilarity = 0
    for sent2 in excel_Q_list:
        similarity = 0
        sent2_cut = list(str(sent2))
        set2 = set(sent2_cut)
        #calculate text similarity
        similarity = difflib.SequenceMatcher(None, sent1_cut, sent2_cut).quick_ratio()
        # find max sim
        if similarity > MAXsimilarity:
            MAXsimilarity = similarity
    correct_excel_Que_list = []
    correct_excel_ANS_list = []
    # save MAX sim to list

    for sent2index in range(len(excel_Q_list)):
        similarity = 0
        sent2_cut = list(str(excel_Q_list[sent2index]))
        set2 = set(sent2_cut)
        #calculate text similarity
        similarity = difflib.SequenceMatcher(None, sent1_cut, sent2_cut).quick_ratio()
        # find max sim
        if similarity == MAXsimilarity:
            correct_excel_Que_list.append(excel_Q_list[sent2index])
            correct_excel_ANS_list.append(excel_ANS_list[sent2index])
    
            # print(sent2index)
            # print(excel_Q_list[sent2index])
            # print(excel_ANS_list[sent2index])

    text = pytesseract.image_to_string(Ans1,config=confignow)
    text1 = text_error_handle(text)
    # print(text)
    text = pytesseract.image_to_string(Ans2,config=confignow)
    text2 = text_error_handle(text)
    # print(text)
    text = pytesseract.image_to_string(Ans3,config=confignow)
    text3 = text_error_handle(text)
    # print(text)
    ORC_YYS_ANS_list = [text1,text2,text3]
    print("ORC YYS Ans list:",ORC_YYS_ANS_list)


    if 1:
        MAX_similarity=0
        CORRECT_INDEX = 0
        # print(correct_excel_ANS_list)
        for correct_ANS_item in correct_excel_ANS_list:
            for ORC_ANS_item in ORC_YYS_ANS_list:
                sent1_cut = set(list(str(correct_ANS_item)))
                sent2_cut = set(list(str(ORC_ANS_item)))
                similarity = difflib.SequenceMatcher(None, sent1_cut, sent2_cut).quick_ratio()
                if similarity > MAX_similarity:
                    MAX_similarity = similarity
        for correct_ANS_item in correct_excel_ANS_list:
            for ORC_ANS_item in ORC_YYS_ANS_list:
                sent1_cut = set(list(str(correct_ANS_item)))
                sent2_cut = set(list(str(ORC_ANS_item)))
                similarity = difflib.SequenceMatcher(None, sent1_cut, sent2_cut).quick_ratio()
                if similarity == MAX_similarity:
                    CORRECT_INDEX = ORC_YYS_ANS_list.index(ORC_ANS_item)
                    print("#### YYS Ans is ",CORRECT_INDEX,ORC_ANS_item)

    
        Ans1_click_Loc_x = 150 + FMMX_Loc_x + 30
        Ans1_click_Loc_y = 15+FMMX_Loc_y + 155
        Ans2_click_Loc_x = 150 + FMMX_Loc_x + 30
        Ans2_click_Loc_y = 15+FMMX_Loc_y + 225
        Ans3_click_Loc_x = 150 + FMMX_Loc_x + 30
        Ans3_click_Loc_y = 15+FMMX_Loc_y + 285

        #draw circle for Ans1_click_Loc_x in screenshot
        # cv2.circle(screenshot,(Ans1_click_Loc_x,Ans1_click_Loc_y), 10, (0,0,255), -1)
        # cv2.circle(screenshot,(Ans2_click_Loc_x,Ans2_click_Loc_y), 10, (0,0,255), -1)
        # cv2.circle(screenshot,(Ans3_click_Loc_x,Ans3_click_Loc_y), 10, (0,0,255), -1)
        Ans_click_Loc_x = [Ans1_click_Loc_x,Ans2_click_Loc_x,Ans3_click_Loc_x]
        Ans_click_Loc_y = [Ans1_click_Loc_y,Ans2_click_Loc_y,Ans3_click_Loc_y]
        # cv2.circle(screenshot,(Ans_click_Loc_x[CORRECT_INDEX],Ans_click_Loc_y[CORRECT_INDEX]), 10, (0,0,255), -1)
        #click Ans
        BASE.click_point(device,Ans_click_Loc_x[CORRECT_INDEX],Ans_click_Loc_y[CORRECT_INDEX])
        time.sleep(1)
        x2 = 0
        y2 = 0
        x2,y2 = BASE.wait_picture(device,YYS_QuDeJiangLi,th_hold=0.95,wait_time=10)
        if x2 or y2:
            x2 = x2+300
            y2 = y2-50
            BASE.click_point(device,x2,y2)
        else:
            # save YYS_Q_text and ORC_YYS_ANS_list to "./pic/FM/error.txt" with UTF8
            with open("./pic/FM/error.txt","a",encoding='utf-8') as f:
                f.write(str(YYS_Q_text)+str(ORC_YYS_ANS_list)+"\n")
                f.write("\n")


    # cv2.imshow("SS",SS)
    # cv2.waitKey(0)